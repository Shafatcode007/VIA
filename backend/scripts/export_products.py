"""Export the grocery product catalog to an Excel workbook.

NECESSITY: The frontend shows products without images. Export every product
with an empty image_url column so images can be collected via web scraping
and the URLs pasted back into the sheet (or used in an UPDATE script).
LOGIC: Reads products via the app's async session and joins seller names.
USAGE:  backend\\.venv\\Scripts\\python.exe backend\\scripts\\export_products.py
"""

import asyncio
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from sqlalchemy import select

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.core.database import AsyncSessionLocal, engine
from app.models.product import Product
from app.models.seller import Seller


HEADERS = [
    "id",
    "name",
    "description",
    "price",
    "unit",
    "stock_quantity",
    "seller_name",
    "seller_id",
    "is_available",
    "image_url",
]

OUTPUT_FILENAME = "products_catalog.xlsx"


async def export() -> None:
    async with AsyncSessionLocal() as session:
        stmt = (
            select(
                Product.id,
                Product.name,
                Product.description,
                Product.price,
                Product.unit,
                Product.stock_quantity,
                Seller.name.label("seller_name"),
                Product.seller_id,
                Product.is_available,
                Product.image_url,
            )
            .join(Seller, Seller.id == Product.seller_id, isouter=True)
            .order_by(Product.is_available.desc(), Product.name.asc())
        )
        rows = (await session.execute(stmt)).all()
        await engine.dispose()

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    header_fill = PatternFill(start_color="4DBE55", end_color="4DBE55", fill_type="solid")
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for row, product in enumerate(rows, start=2):
        for col, value in enumerate(product, start=1):
            ws.cell(row=row, column=col, value=value)

    widths = [6, 40, 50, 10, 10, 14, 28, 10, 12, 60]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    out_path = Path(__file__).resolve().parents[1] / OUTPUT_FILENAME
    wb.save(out_path)
    print(f"Exported {len(rows)} products -> {out_path}")


if __name__ == "__main__":
    asyncio.run(export())