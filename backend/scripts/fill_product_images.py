# backend/scripts/fill_product_images.py
"""
Fill the empty image_url column of backend/products_catalog.xlsx.

Strategy (license-clean, offline-safe, re-runnable):
1. Normalize each product name to a canonical grocery item.
2. Fetch one representative CC-licensed image per canonical item from the
   Wikimedia Commons API (no retailer scraping, no watermarks).
3. Download the 480px thumbnail once into via-app/public/products/ so the
   running frontend (via-app on :3000) can serve it offline as /products/*.
4. Write the LOCAL path (/products/<slug>.<ext>) into image_url so the
   demo works fully offline and never depends on hotlinking.
5. Emit product_image_urls.json + update_image_urls.sql for the DB import.

Run anytime (also after re-running export_products.py):
    backend\.venv\Scripts\python.exe backend\scripts\fill_product_images.py
"""

from __future__ import annotations

import json
import re
import sys
import time
import urllib.parse
import urllib.request
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl required: backend\.venv\Scripts\python.exe -m pip install openpyxl")

PROJECT_ROOT = Path(__file__).resolve().parents[2]
XLSX_PATH = PROJECT_ROOT / "backend" / "products_catalog.xlsx"
PUBLIC_DIR = PROJECT_ROOT / "via-app" / "public" / "products"
SCRIPT_DIR = Path(__file__).resolve().parent
OUT_JSON = SCRIPT_DIR / "product_image_urls.json"
OUT_SQL = SCRIPT_DIR / "update_image_urls.sql"

USER_AGENT = "VIA-CSE327-UniversityDemo/1.0"

# canonical key -> (Commons search terms, file slug)
CANONICAL_ITEMS: dict[str, tuple[str, str]] = {
    "rice": ("white rice grains bowl", "rice"),
    "onion": ("red onions", "onion"),
    "potato": ("potatoes", "potato"),
    "tomato": ("ripe red tomatoes", "tomato"),
    "garlic": ("garlic bulbs", "garlic"),
    "chili": ("green chili peppers", "green-chili"),
    "ginger": ("ginger root", "ginger"),
    "beef": ("raw beef meat", "beef"),
    "chicken": ("raw chicken meat", "chicken"),
    "egg": ("brown chicken eggs carton", "egg"),
    "milk": ("milk bottle", "milk"),
    "cooking oil": ("cooking oil bottle", "cooking-oil"),
    "lentil": ("red lentils masoor", "lentil"),
    "sugar": ("white sugar crystals", "sugar"),
    "salt": ("white salt bowl", "salt"),
    "flour": ("wheat flour bowl", "flour"),
"tea": ("dried black tea leaves", "tea"),
    "biscuit": ("tea biscuits cookies", "biscuit"),
    "noodles": ("instant noodles", "noodles"),
    "fish": ("rohu fish", "fish"),
    "mutton": ("lamb meat", "mutton"),
    "green chili": ("green chili peppers", "green-chili"),
}

ALIASES: list[tuple[str, str]] = [
    ("miniket", "rice"), ("rice", "rice"),
    ("onion", "onion"), ("potato", "potato"), ("tomato", "tomato"),
    ("garlic", "garlic"), ("chili", "green chili"), ("chilli", "green chili"),
    ("ginger", "ginger"), ("beef", "beef"), ("chicken", "chicken"),
    ("egg", "egg"), ("milk", "milk"), ("oil", "cooking oil"),
    ("masoor", "lentil"), ("dal", "lentil"), ("lentil", "lentil"),
    ("sugar", "sugar"), ("salt", "salt"), ("atta", "flour"), ("flour", "flour"),
    ("tea", "tea"), ("biscuit", "biscuit"), ("noodl", "noodles"),
    ("rui", "fish"), ("rohu", "fish"), ("fish", "fish"),
    ("mutton", "mutton"),
]

# Hand-picked overrides that must never be re-scraped.
# Files must exist in via-app/public/products (or via-app/public/products).
MANUAL_OVERRIDES: dict[str, str] = {
    "soybean-oil": "/products/cooking-oil.svg",
    "biscuit": "/products/biscuit.svg",
    "cooking-oil": "/products/cooking-oil.svg",
}

SIZE_RE = re.compile(r"\d+(?:\.\d+)?\s*(?:kg|g|ml|l|dozen|pcs|pieces|poya)\b", re.I)


def canonical_key(name: str) -> str | None:
    cleaned = SIZE_RE.sub(" ", name.lower())
    cleaned = re.sub(r"\b(fresh|local|premium|pure|whole)\b", " ", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    for alias, key in ALIASES:
        if alias in cleaned:
            return key
    return None


def column_index(header: list, prefix: str) -> int:
    for position, value in enumerate(header):
        if value and str(value).strip().lower().startswith(prefix):
            return position
    raise KeyError(f"column starting with '{prefix}' not found in {header}")


def commons_thumb_url(search: str) -> str | None:
    """First bitmap match on Wikimedia Commons, 480px thumbnail URL."""
    query = urllib.parse.urlencode({
        "action": "query", "generator": "search",
        "gsrsearch": f"filetype:bitmap {search}", "gsrnamespace": "6", "gsrlimit": "1",
        "prop": "imageinfo", "iiprop": "url|mime", "iiurlwidth": "480", "format": "json",
    })
    request = urllib.request.Request(
        f"https://commons.wikimedia.org/w/api.php?{query}",
        headers={"User-Agent": USER_AGENT},
    )
    with urllib.request.urlopen(request, timeout=30) as response:
        data = json.load(response)
    for page in ((data.get("query") or {}).get("pages") or {}).values():
        info = (page.get("imageinfo") or [{}])[0]
        if str(info.get("mime", "")).startswith("image/"):
            return info.get("thumburl") or info.get("url")
    return None


def download(url: str, destination: Path) -> None:
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(request, timeout=60) as response:
        destination.write_bytes(response.read())


def write_placeholder(slug: str, label: str, destination: Path) -> None:
    """Deterministic offline fallback tile so no row ever stays empty."""
    destination.write_text(
        "<svg xmlns='http://www.w3.org/2000/svg' width='480' height='360'>"
        "<rect width='100%' height='100%' fill='#edf7ee'/>"
        f"<text x='50%' y='52%' font-family='Arial' font-size='140' fill='#4DBE55' "
        f"text-anchor='middle' dominant-baseline='middle'>{label[:1].upper()}</text>"
        f"<text x='50%' y='86%' font-family='Arial' font-size='28' fill='#71776D' "
        f"text-anchor='middle'>{label.title()}</text></svg>",
        encoding="utf-8",
    )


def main() -> int:
    if not XLSX_PATH.exists():
        print(f"[error] workbook not found: {XLSX_PATH}")
        return 1
    PUBLIC_DIR.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(XLSX_PATH)
    sheet = workbook.active
    header = [cell.value for cell in sheet[1]]
    name_col = column_index(header, "name")
    image_col = column_index(header, "image_url")

    key_to_url: dict[str, str] = {}
    name_to_url: dict[str, str] = {}
    filled = unmatched = 0

    for row in sheet.iter_rows(min_row=2):
        name = row[name_col].value
        if not name:
            continue
        name = str(name)
        key = canonical_key(name)
        if key is None:
            unmatched += 1
            print(f"[warn] no canonical match: {name}")
            continue
        if key not in key_to_url:
            # Check manual overrides first
            if key in MANUAL_OVERRIDES:
                local_path = MANUAL_OVERRIDES[key]
                print(f"[override] {key} -> {local_path}")
            else:
                search, slug = CANONICAL_ITEMS[key]
                local_path: str | None = None
                existing = next(PUBLIC_DIR.glob(f"{slug}.*"), None)
                if existing is not None:
                    local_path = f"/products/{existing.name}"
                else:
                    try:
                        remote = commons_thumb_url(search)
                        time.sleep(1)
                    except Exception as exc:
                        remote = None
                        print(f"[warn] commons lookup failed for {key}: {exc}")
                    if remote:
                        suffix = remote.rsplit(".", 1)[-1].split("?")[0].lower()
                        ext = suffix if suffix in {"jpg", "jpeg", "png", "webp"} else "jpg"
                        target = PUBLIC_DIR / f"{slug}.{ext}"
                        try:
                            download(remote, target)
                            local_path = f"/products/{target.name}"
                        except Exception as exc:
                            print(f"[warn] download failed for {key}: {exc}")
                    if local_path is None:
                        target = PUBLIC_DIR / f"{slug}.svg"
                        write_placeholder(slug, key, target)
                        local_path = f"/products/{target.name}"
                key_to_url[key] = local_path
                print(f"[ok] {key} -> {local_path}")
            row[image_col].value = key_to_url[key]
            name_to_url[name] = key_to_url[key]
            filled += 1

    workbook.save(XLSX_PATH)
    OUT_JSON.write_text(json.dumps({"by_name": name_to_url, "by_key": key_to_url}, indent=2), encoding="utf-8")

    lines = ["BEGIN;"]
    for name, url in name_to_url.items():
        lines.append(f"UPDATE products SET image_url = '{url}' WHERE name = '{name.replace(chr(39), chr(39) * 2)}';")
    lines.append("COMMIT;")
    OUT_SQL.write_text("\n".join(lines) + "\n", encoding="utf-8")

    print(f"\nDone. Filled {filled} rows, {unmatched} unmatched.")
    print(f"Mapping: {OUT_JSON}\nDB import SQL: {OUT_SQL}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())